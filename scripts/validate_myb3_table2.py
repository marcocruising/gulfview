#!/usr/bin/env python3
"""
Post-parse QA for USGS myb3 Table 2.

Run from repo root:
  uv run python scripts/validate_myb3_table2.py

Checks:
  1) No literal ``do.`` left in owner_operator / location / capacity_raw (ditto expansion).
  2) Rows whose commodity cell is ``Do.``: commodity_leaf_resolved matches an independent
     upward scan in column A (skipping chrome, Continued, empty, ditto).
  3) Rows whose Excel commodity cell is ``Do.``: ``commodity_cell_raw`` equals ``commodity_leaf_resolved``
     (resolved USGS label, not NULL).

Does not validate owner/location ditto against “cell above” in Excel — wrapped rows put
fragments in the owner column, while the loader uses last explicit value on block starts
(USGS table convention).
"""

from __future__ import annotations

import sys
from pathlib import Path

_ROOT = Path(__file__).resolve().parent.parent
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

import openpyxl  # noqa: E402

import loaders.load_usgs as M  # noqa: E402

USGS = _ROOT / "data" / "usgs"


def _cell(row: list[object], j: int) -> str:
    if j >= len(row):
        return ""
    return M._norm_text(row[j])


def _prev_commodity_leaf(matrix: list[list], header_row: int, start_row: int) -> str:
    """Previous non-ditto commodity text; skip chrome and bare 'Continued'."""
    for r in range(start_row - 1, header_row, -1):
        row = matrix[r]
        if not row:
            continue
        c0 = _cell(row, 0)
        if not c0:
            continue
        if M._is_t2_chrome_row(c0):
            continue
        if M._is_ditto_token(c0):
            continue
        if M._is_table2_wrap_continued_cell(c0):
            continue
        low = c0.lower()
        if low == "continued":
            continue
        if low == "content":
            continue
        return c0
    return ""


def main() -> int:
    paths = sorted(USGS.glob("myb3*.xlsx"))
    if not paths:
        print(f"No myb3*.xlsx under {USGS}", file=sys.stderr)
        return 1

    inv_ditto_left = 0
    inv_commodity = 0
    inv_cell_raw = 0
    inv_idem = 0
    samples: list[str] = []
    total_blocks = 0

    for path in paths:
        meta = M._parse_myb3_filename(path.name)
        if not meta:
            continue
        refy, iso3 = meta
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        s2 = M._find_workbook_sheet(wb.sheetnames, r"table\s*2")
        if not s2:
            wb.close()
            print(f"{path.name}: no Table 2", file=sys.stderr)
            return 1
        matrix = [list(r) for r in wb[s2].iter_rows(values_only=True)]
        wb.close()

        det = M._t2_detect_columns(matrix)
        if not det:
            print(f"{path.name}: column detection failed", file=sys.stderr)
            return 1
        hr, io, il, icap = det
        assert io is not None and il is not None and icap is not None

        blocks = M.parse_myb3_table2(
            matrix,
            source_file=path.name,
            sheet_name=s2,
            country_iso3=iso3,
            reference_year=refy,
            pulled_at="validate",
        )
        total_blocks += len(blocks)

        blocks2 = M.parse_myb3_table2(
            matrix,
            source_file=path.name,
            sheet_name=s2,
            country_iso3=iso3,
            reference_year=refy,
            pulled_at="validate2",
        )
        fps = {b["record_fingerprint"] for b in blocks}
        fps2 = {b["record_fingerprint"] for b in blocks2}
        if fps != fps2 or len(blocks) != len(blocks2):
            samples.append(f"{path.name}: idempotency fail (re-parse changed fingerprints or count)")
            inv_idem += 1

        for b in blocks:
            rs = b["excel_row_start"]
            row = matrix[rs]
            raw_c0 = _cell(row, 0)

            for fld, v in (
                ("owner_operator", b.get("owner_operator") or ""),
                ("location", b.get("location") or ""),
                ("capacity_raw", b.get("capacity_raw") or ""),
            ):
                if M._is_ditto_token(v):
                    inv_ditto_left += 1
                    samples.append(f"{path.name} r{rs} {fld} still ditto: {v!r}")

            # Commodity-column Do.: must match previous explicit commodity (skip Continued / chrome).
            if M._is_ditto_token(raw_c0):
                exp = _prev_commodity_leaf(matrix, hr, rs)
                got = b.get("commodity_leaf_resolved") or ""
                if exp != got:
                    inv_commodity += 1
                    samples.append(
                        f"{path.name} r{rs} commodity Do.: want {exp[:55]!r} got {got[:55]!r}"
                    )
                cr = b.get("commodity_cell_raw") or ""
                got_leaf = b.get("commodity_leaf_resolved") or ""
                if cr != got_leaf or not cr:
                    inv_cell_raw += 1
                    samples.append(
                        f"{path.name} r{rs} Do. row: cell_raw {cr[:40]!r} should match leaf {got_leaf[:40]!r}"
                    )

    print(f"Files: {len(paths)}  Table 2 blocks: {total_blocks}")
    print(
        "Checks: unresolved do. o/l/cap | commodity Do. replay | cell_raw=leaf if Do. | idempotent re-parse"
    )
    print(f"        {inv_ditto_left} | {inv_commodity} | {inv_cell_raw} | {inv_idem}")
    if samples:
        print("\nFirst mismatches / issues:")
        for line in samples[:25]:
            print(" ", line)
        if len(samples) > 25:
            print(f"  ... and {len(samples) - 25} more")
        return 1
    print("OK — all independent replay checks passed.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
