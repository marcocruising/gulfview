# ============================================================
# SCRIPT:  load_gem_xlsx.py
# SOURCE:  Global Energy Monitor — Excel tracker downloads (manual)
# URL:     https://www.globalenergymonitor.org/
# API KEY: not required
# WRITES:  gem_tracker_rows
# REFRESH: Re-download xlsx and re-run; each sheet is delete-then-insert for that source_file+sheet.
# NOTES:   Default bundle = industrial/plant trackers + GGIT/GOIT pipelines/LNG (see DEFAULT_WORKBOOKS).
#          Uses openpyxl read-only streaming. Skips About/Metadata unless you pass --include-meta-sheets.
# ============================================================

from __future__ import annotations

import argparse
import sys
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl

_ROOT = Path(__file__).resolve().parent.parent
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

from utils.pipeline_logger import finish_run, start_run
from utils.supabase_client import get_client

SCRIPT_NAME = "load_gem_xlsx"
SOURCE_LABEL = "Global Energy Monitor (GEM) Excel trackers"
GEM_DIR = _ROOT / "data" / "globalenergymonitor"

# Default: user-requested subset (data sheets only).
DEFAULT_WORKBOOKS: dict[str, list[str]] = {
    "Global-Cement-and-Concrete-Tracker_July-2025.xlsx": ["Plant Data"],
    "Global-Iron-Ore-Mines-Tracker-August-2025-V1.xlsx": ["Main Data"],
    "Plant-level-data-Global-Chemicals-Inventory-November-2025-V1.xlsx": ["Plant data"],
    "Plant-level-data-Global-Iron-and-Steel-Tracker-March-2026-V1.xlsx": [
        "Plant data",
        "Plant capacities and status",
        "Plant production",
    ],
    "GEM-GOIT-Oil-NGL-Pipelines-2025-03.xlsx": ["Pipelines"],
    "GEM-GGIT-LNG-Terminals-2025-09.xlsx": ["LNG Terminals"],
    "GEM-GGIT-Gas-Pipelines-2025-11.xlsx": ["Pipelines"],
    "Global-Integrated-Power-March-2026-II.xlsx": [
        "Power facilities",
        "Regions, area, and countries",
    ],
}

SKIP_SHEETS = frozenset({"About", "Metadata"})
INSERT_BATCH = 500


def _cell_to_json(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, float) and value != value:
        return None
    if isinstance(value, (int, bool)):
        return value
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    s = str(value).strip()
    return s if s else None


def _normalize_headers(raw: list[Any]) -> list[str]:
    seen: dict[str, int] = {}
    out: list[str] = []
    for i, c in enumerate(raw):
        s = ("" if c is None else str(c)).strip()
        if not s:
            s = f"_col_{i}"
        key = s
        n = seen.get(key, 0)
        seen[key] = n + 1
        if n:
            s = f"{key}__{n + 1}"
        out.append(s)
    return out


def _row_is_empty(payload: dict[str, Any]) -> bool:
    for v in payload.values():
        if v is not None and v != "":
            return False
    return True


def _load_sheet(
    path: Path,
    sheet_name: str,
    source_file: str,
    pulled_at: str,
    source_tag: str,
) -> tuple[int, list[dict[str, Any]]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet {sheet_name!r} not in {wb.sheetnames}")
        ws = wb[sheet_name]
        it = ws.iter_rows(values_only=True)
        try:
            header_row = next(it)
        except StopIteration:
            return 0, []
        headers = _normalize_headers(list(header_row))
        rows_out: list[dict[str, Any]] = []
        excel_row = 2
        for data_row in it:
            vals = list(data_row[: len(headers)])
            if len(vals) < len(headers):
                vals.extend([None] * (len(headers) - len(vals)))
            payload = {h: _cell_to_json(v) for h, v in zip(headers, vals, strict=True)}
            if not _row_is_empty(payload):
                rows_out.append(
                    {
                        "source_file": source_file,
                        "sheet_name": sheet_name,
                        "excel_row_1based": excel_row,
                        "payload": payload,
                        "source": source_tag,
                        "pulled_at": pulled_at,
                    }
                )
            excel_row += 1
        return len(rows_out), rows_out
    finally:
        wb.close()


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Load GEM .xlsx data sheets into gem_tracker_rows (default: industrial + GGIT/GOIT bundle)."
    )
    ap.add_argument(
        "--file",
        action="append",
        dest="files",
        metavar="NAME.xlsx",
        help="Load one workbook (under data/globalenergymonitor/). Repeatable. "
        "Without --sheets, loads every sheet except About and Metadata.",
    )
    ap.add_argument(
        "--sheets",
        type=str,
        help="Comma-separated sheet names (only with a single --file).",
    )
    ap.add_argument(
        "--include-meta-sheets",
        action="store_true",
        help="When using --file without --sheets, also load About and Metadata sheets.",
    )
    ap.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse files and print counts; do not touch Supabase.",
    )
    args = ap.parse_args()

    if args.files and len(args.files) > 1 and args.sheets:
        print("--sheets is only supported with a single --file", file=sys.stderr)
        return 1

    if args.dry_run:
        client = None
        run_id = None
    else:
        try:
            client = get_client()
        except RuntimeError as e:
            print(e, file=sys.stderr)
            return 1
        params: dict[str, Any] = {
            "mode": "custom" if args.files else "default_bundle",
            "files": args.files,
            "sheets": args.sheets,
            "include_meta": args.include_meta_sheets,
        }
        try:
            run_id = start_run(client, SCRIPT_NAME, SOURCE_LABEL, params)
        except Exception as e:
            print(f"Could not log pipeline run: {e}", file=sys.stderr)
            return 1

    pulled_at = datetime.now(timezone.utc).isoformat()
    source_tag = SCRIPT_NAME
    total_written = 0

    jobs: list[tuple[Path, str]] = []

    if args.files:
        for fn in args.files:
            p = GEM_DIR / fn
            if not p.is_file():
                msg = f"File not found: {p}"
                if client and run_id is not None:
                    finish_run(client, run_id, 0, "error", msg)
                print(msg, file=sys.stderr)
                return 1
            if args.sheets:
                for sn in [s.strip() for s in args.sheets.split(",") if s.strip()]:
                    jobs.append((p, sn))
            else:
                wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
                try:
                    names = list(wb.sheetnames)
                finally:
                    wb.close()
                for sn in names:
                    if not args.include_meta_sheets and sn in SKIP_SHEETS:
                        continue
                    jobs.append((p, sn))
    else:
        for fn, sheets in DEFAULT_WORKBOOKS.items():
            p = GEM_DIR / fn
            if not p.is_file():
                msg = f"Default bundle: missing {p.name} (place under data/globalenergymonitor/)"
                if client and run_id is not None:
                    finish_run(client, run_id, 0, "error", msg)
                print(msg, file=sys.stderr)
                return 1
            for sn in sheets:
                jobs.append((p, sn))

    try:
        for path, sheet_name in jobs:
            n, rows = _load_sheet(
                path, sheet_name, path.name, pulled_at, source_tag
            )
            print(f"{path.name} / {sheet_name!r}: {n} data rows")
            if args.dry_run:
                total_written += n
                continue
            assert client is not None
            client.table("gem_tracker_rows").delete().eq("source_file", path.name).eq(
                "sheet_name", sheet_name
            ).execute()
            for i in range(0, len(rows), INSERT_BATCH):
                batch = rows[i : i + INSERT_BATCH]
                client.table("gem_tracker_rows").insert(batch).execute()
                total_written += len(batch)

        if client and run_id is not None:
            finish_run(client, run_id, total_written, "success", None)
        if args.dry_run:
            print(f"Dry run: would insert {total_written} rows across {len(jobs)} sheet(s).")
        else:
            print(f"Inserted {total_written} rows into gem_tracker_rows.")
        return 0
    except Exception as e:
        err = f"{type(e).__name__}: {e}"
        if client and run_id is not None:
            finish_run(client, run_id, total_written, "error", err)
        print(err, file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
