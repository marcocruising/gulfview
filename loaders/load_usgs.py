# ============================================================
# SCRIPT:  load_usgs.py
# SOURCE:  USGS Mineral Commodity Summaries (MCS) CSV + myb3 xlsx
# URL:     https://www.usgs.gov/centers/national-minerals-information-center
# API KEY: not required (manual download)
# WRITES:  usgs_mineral_statistics (mcs); usgs_myb3_production + usgs_country_mineral_facilities (facilities)
# REFRESH: annual when USGS publishes new MCS tables / yearbooks
# NOTES:   MCS CSV under data/usgs/ (cp1252). myb3-*.xlsx: Table 1 melt + Table 2 merged blocks.
#
#   myb3 Table 1 (parse_myb3_table1): sheet ``^table\\s*1$``; first row with col A starting with
#   ``Commodity`` = header; year columns melted; section rows (METALS / INDUSTRIAL MINERALS) + colon
#   hierarchy in commodity_path; footnotes r/e in spacer columns.
#
#   myb3 Table 2 (parse_myb3_table2) — USGS conventions for future country files:
#   - Do./do. in commodity column → commodity_leaf_resolved = last explicit label; commodity_cell_raw stores that
#     same resolved label (not NULL) so every row is fully labeled in the UI.
#   - Do./do. in owner / location / capacity → repeat last explicit value in that column (block starts + flush
#     writes merged O/L/C so ditto after wrapped rows repeats the full merged text).
#   - "Continued" / "—Continued" in column A (no trailing ":") → blank commodity (wrap). Section lines
#     like ``Iron and steel—Continued:`` keep the colon and use the colon-header path.
#   - Stray ``content`` in A with empty owner + location text → continuation (Iran chromite wrap).
#   - facility_path = normalized colon stack + commodity leaf (leaf appended even when stack was cleared).
#   - Skip chrome: standalone Commodity, TABLE 2—Continued, country title, footnote line; skip empty-A rows
#     that only repeat subheader fragments (Major operating companies…).
#   - New files: name myb3-{year}-{slug}.xlsx and add slug → ISO3 in MYB3_SLUG_TO_ISO3.
#   - Before upsert, existing rows with the same source_file are deleted (avoids orphan rows when
#     record_fingerprint changes).
#   - QA: uv run python scripts/validate_myb3_table2.py
#
#   Adding another country (checklist):
#   1) Save as data/usgs/myb3-{YEAR}-{slug}.xlsx (slug = lowercase hyphenated, e.g. kuwait).
#   2) Add slug → ISO3 in MYB3_SLUG_TO_ISO3 (unmapped files are skipped; pipeline_runs may be partial).
#   3) Keep standard three-sheet layout: Text (ignored), Table 1 / Table 1, Table 2 / Table 2.
#   4) Run: uv run python scripts/validate_myb3_table2.py && uv run python loaders/load_usgs.py facilities
#
#   Residual risks (rare in current files; extend code if you hit them):
#   - Col A is only a hierarchy label with no owner/loc/cap on that line, then Do. on the next row:
#     last commodity may not update until we add explicit handling.
#   - Table 2 header text changes so much that _t2_detect_columns cannot find owner/loc/cap columns.
#   - A top-level product word not in _T2_ROOT_FIRST_WORDS may leave a deeper facility_stack in place;
#     add the word to the frozenset if grouping looks wrong.
# ============================================================

from __future__ import annotations

# --- CONFIGURATION — edit these values before running --------
MCS_GLOB = "MCS*_Commodities_Data.csv"
UPSERT_BATCH = 500
# -------------------------------------------------------------

import argparse
import hashlib
import re
import sys
import unicodedata
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
import pycountry

_ROOT = Path(__file__).resolve().parent.parent
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

USGS_DIR = _ROOT / "data" / "usgs"

from utils.pipeline_logger import finish_run, start_run
from utils.supabase_client import get_client

SCRIPT_NAME = "load_usgs"
SOURCE_LABEL = "USGS Mineral Commodity Summaries (MCS) CSV"
SOURCE_LABEL_MYB3 = "USGS Minerals Yearbook myb3 country xlsx (Table 1 + Table 2)"

MYB3_FILENAME_RE = re.compile(r"^myb3-(\d{4})-(.+)\.xlsx$", re.I)
# Add an entry for each new regional workbook slug (hyphenated, lower case).
MYB3_SLUG_TO_ISO3: dict[str, str] = {
    "bahrain": "BHR",
    "iraq": "IRQ",
    "oman": "OMN",
    "qatar": "QAT",
    "united-arab-emirates": "ARE",
    "iran": "IRN",
    "saudi-arabia": "SAU",
}

# Aggregates and multi-country rows — no single ISO3.
COUNTRY_NAME_TO_ISO3: dict[str, str | None] = {
    "World total": None,
    "Other countries": None,
    "China, Germany, and Russia": None,
    "United States": "USA",
    "United Kingdom": "GBR",
    "Korea, Republic of": "KOR",
    "Korea, North": "PRK",
    "Burma": "MMR",
    "Congo (Kinshasa)": "COD",
    "The Bahamas": "BHS",
    "Taiwan": "TWN",
    "Russia": "RUS",
    "Iran": "IRN",
    "Vietnam": "VNM",
    "Laos": "LAO",
    "Côte d'Ivoire": "CIV",
}

EXPECTED_MCS_COLS = [
    "MCS chapter",
    "Section",
    "Commodity",
    "Country",
    "Statistics",
    "Statistics_detail",
    "Unit",
    "Year",
    "Value",
    "Notes",
    "Is critical mineral 2025",
    "Other notes",
]


def _norm_text(x: object) -> str:
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return ""
    s = str(x).strip()
    return unicodedata.normalize("NFKC", s)


def _country_to_iso3(name: str) -> str | None:
    key = _norm_text(name)
    if not key:
        return None
    key_alt = key.replace("\u0092", "'").replace("\u2019", "'")
    if key in COUNTRY_NAME_TO_ISO3:
        return COUNTRY_NAME_TO_ISO3[key]
    if key_alt in COUNTRY_NAME_TO_ISO3:
        return COUNTRY_NAME_TO_ISO3[key_alt]
    if "ivoire" in key.lower():
        return "CIV"
    try:
        c = pycountry.countries.get(name=key)
        if c:
            return c.alpha_3
    except (LookupError, KeyError, AttributeError, TypeError):
        pass
    try:
        matches = pycountry.countries.search_fuzzy(key)
        if matches:
            return matches[0].alpha_3
    except (LookupError, AttributeError):
        pass
    return None


def _parse_critical(val: object) -> bool | None:
    s = _norm_text(val).lower()
    if s in ("yes", "y", "true", "1"):
        return True
    if s in ("no", "n", "false", "0"):
        return False
    return None


def _parse_value_numeric(raw: str) -> float | None:
    s = _norm_text(raw)
    if not s:
        return None
    cleaned = s.replace(",", "").strip()
    if re.fullmatch(r"-?\d+(\.\d+)?", cleaned):
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def _parse_year(val: object) -> tuple[int | None, str]:
    s = _norm_text(val)
    if not s:
        return None, ""
    if re.fullmatch(r"\d{4}", s):
        return int(s), s
    m = re.match(r"(\d{4})\s*[-–]\s*(\d{2}|\d{4})", s)
    if m:
        return int(m.group(1)), s
    return None, s


def _record_fingerprint(
    chapter: str,
    section: str,
    commodity: str,
    country: str,
    statistics: str,
    detail: str,
    unit: str,
    year_as_reported: str,
    value_raw: str,
    notes: str,
) -> str:
    parts = [
        chapter,
        section,
        commodity,
        country,
        statistics,
        detail,
        unit,
        year_as_reported,
        value_raw,
        notes,
    ]
    joined = "\x1e".join(_norm_text(p) for p in parts)
    return hashlib.sha256(joined.encode("utf-8")).hexdigest()


def _discover_mcs_files() -> list[Path]:
    if not USGS_DIR.is_dir():
        return []
    return sorted(USGS_DIR.glob(MCS_GLOB))


def _mcs_dataframe_to_rows(df: pd.DataFrame, source_file: str, pulled_at: str) -> list[dict[str, Any]]:
    cols = [c.strip() for c in df.columns]
    df = df.copy()
    df.columns = cols
    missing = [c for c in EXPECTED_MCS_COLS if c not in df.columns]
    if missing:
        raise ValueError(f"Missing MCS columns: {missing}")

    rows: list[dict[str, Any]] = []
    for _, r in df.iterrows():
        chapter = _norm_text(r["MCS chapter"])
        section = _norm_text(r["Section"])
        commodity = _norm_text(r["Commodity"])
        country_name = _norm_text(r["Country"])
        statistics = _norm_text(r["Statistics"])
        detail = _norm_text(r["Statistics_detail"])
        unit = _norm_text(r["Unit"])
        notes = _norm_text(r["Notes"])
        other = _norm_text(r["Other notes"])
        value_raw = _norm_text(r["Value"])
        data_year, year_as_reported = _parse_year(r["Year"])
        if data_year is None:
            continue

        fp = _record_fingerprint(
            chapter,
            section,
            commodity,
            country_name,
            statistics,
            detail,
            unit,
            year_as_reported,
            value_raw,
            notes,
        )
        iso3 = _country_to_iso3(country_name)
        crit = _parse_critical(r["Is critical mineral 2025"])
        num = _parse_value_numeric(value_raw)

        rows.append(
            {
                "record_fingerprint": fp,
                "mcs_chapter": chapter,
                "section": section,
                "commodity": commodity,
                "country_name": country_name,
                "country_iso3": iso3,
                "statistics": statistics,
                "statistics_detail": detail,
                "unit": unit,
                "data_year": data_year,
                "year_as_reported": year_as_reported,
                "value_numeric": num,
                "value_raw": value_raw if value_raw else None,
                "notes": notes if notes else None,
                "other_notes": other if other else None,
                "is_critical_mineral_2025": crit,
                "source_file": source_file,
                "source": SCRIPT_NAME,
                "pulled_at": pulled_at,
            }
        )
    return rows


def _find_workbook_sheet(sheetnames: list[str], pattern: str) -> str | None:
    rx = re.compile(rf"^{pattern}$", re.I)
    for n in sheetnames:
        if rx.match(n.strip()):
            return n
    return None


def _sheet_to_matrix(ws: Any, max_row: int = 4000, max_col: int = 32) -> list[list[Any]]:
    out: list[list[Any]] = []
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
        out.append(list(row))
    return out


def _parse_myb3_filename(name: str) -> tuple[int, str] | None:
    m = MYB3_FILENAME_RE.match(name)
    if not m:
        return None
    year = int(m.group(1))
    slug = m.group(2).lower().replace("_", "-")
    iso3 = MYB3_SLUG_TO_ISO3.get(slug)
    if not iso3:
        return None
    return year, iso3


def _t1_is_section_all_caps(c0: str) -> bool:
    if not c0 or len(c0) > 90:
        return False
    letters = [ch for ch in c0 if ch.isalpha()]
    if len(letters) < 4:
        return False
    return c0.upper() == c0


def _t1_colon_line_resets_branch(c0: str) -> bool:
    """Heuristic: major commodity ':' lines start a new branch, not under N content / similar."""
    low = c0.lower()
    starters = (
        "stone",
        "sand ",
        "cement",
        "petroleum",
        "natural gas",
        "refinery",
        "mineral fuel",
        "iron and steel",
        "copper:",
        "aluminum:",
        "gold",
        "silver",
        "zinc",
        "lead",
        "nickel",
    )
    return any(low.startswith(s) for s in starters)


def _t1_row_has_year_value(row: list[Any], year_cols: list[tuple[int, int]]) -> bool:
    for yc, _ in year_cols:
        if yc >= len(row):
            continue
        v = _norm_text(row[yc])
        if v and v != "--":
            return True
    return False


def _t1_find_header_and_years(matrix: list[list[Any]]) -> tuple[int, list[tuple[int, int]], str | None] | None:
    unit_context: str | None = None
    for i, row in enumerate(matrix[:25]):
        if row and row[0] is not None:
            t = _norm_text(row[0])
            if "metric" in t.lower() or "ton" in t.lower():
                if "(" in t or "unless" in t.lower():
                    unit_context = t

    header_idx: int | None = None
    for i, row in enumerate(matrix[:45]):
        if not row or row[0] is None:
            continue
        c0 = _norm_text(row[0]).lower()
        if c0.startswith("commodity"):
            header_idx = i
            break
    if header_idx is None:
        return None

    header = matrix[header_idx]
    year_cols: list[tuple[int, int]] = []
    for j, cell in enumerate(header):
        if cell is None:
            continue
        s = _norm_text(cell)
        if re.fullmatch(r"\d{4}", s):
            year_cols.append((j, int(s)))
    if not year_cols:
        return None
    return header_idx, year_cols, unit_context


def _myb3_fp_production(
    source_file: str,
    country_iso3: str,
    commodity_path: str,
    stat_year: int,
    value_raw: str,
    footnote: str,
) -> str:
    blob = "\x1e".join(
        [source_file, country_iso3, commodity_path, str(stat_year), value_raw, footnote]
    )
    return hashlib.sha256(blob.encode("utf-8")).hexdigest()


def parse_myb3_table1(
    matrix: list[list[Any]],
    *,
    source_file: str,
    sheet_name: str,
    country_iso3: str,
    reference_year: int,
    pulled_at: str,
) -> list[dict[str, Any]]:
    found = _t1_find_header_and_years(matrix)
    if not found:
        return []
    header_idx, year_cols, unit_context = found

    major_section: str | None = None
    branch_stack: list[str] = []
    out: list[dict[str, Any]] = []

    for ri in range(header_idx + 1, len(matrix)):
        row = matrix[ri]
        if not row:
            continue
        c0 = _norm_text(row[0])
        if not c0:
            continue
        low = c0.lower()
        if low.startswith("eestimated") or "table includes" in low or low.startswith("1table"):
            break
        if "ditto" in low and "zero" in low and len(c0) < 120:
            break

        has_years = _t1_row_has_year_value(row, year_cols)

        if not has_years:
            if _t1_is_section_all_caps(c0):
                major_section = c0
                branch_stack.clear()
            elif c0.rstrip().endswith(":") or "content:" in low:
                # New ':' heading often starts a fresh branch; long headings (e.g. stone/sand)
                # should not nest under a prior short group (e.g. Nitrogen, N content:).
                if len(c0) > 42 or _t1_colon_line_resets_branch(c0):
                    branch_stack.clear()
                branch_stack.append(c0)
            continue

        path_parts = [p for p in [major_section, *branch_stack, c0] if p]
        # Iraq-style tables: Salt (and similar) rows sit after N-fertilizer children but before the next ':' branch.
        if (
            len(branch_stack) == 1
            and "nitrogen" in branch_stack[0].lower()
            and c0.strip().lower() in ("salt", "sand and gravel, industrial, silica")
        ):
            path_parts = [p for p in [major_section, c0] if p]
        commodity_path = " > ".join(path_parts)

        for yc, stat_year in year_cols:
            if yc >= len(row):
                continue
            v_raw = _norm_text(row[yc])
            if not v_raw or v_raw == "--":
                continue
            fn_col = yc + 1
            foot = _norm_text(row[fn_col]) if fn_col < len(row) else ""
            if foot.lower() in ("", "none"):
                footnote: str | None = None
            else:
                footnote = foot

            fp = _myb3_fp_production(
                source_file, country_iso3, commodity_path, stat_year, v_raw, footnote or ""
            )
            out.append(
                {
                    "record_fingerprint": fp,
                    "country_iso3": country_iso3,
                    "reference_year": reference_year,
                    "commodity_path": commodity_path,
                    "stat_year": stat_year,
                    "value_raw": v_raw,
                    "value_numeric": _parse_value_numeric(v_raw),
                    "footnote": footnote,
                    "unit_context": unit_context,
                    "source_file": source_file,
                    "sheet_name": sheet_name,
                    "source": SCRIPT_NAME,
                    "pulled_at": pulled_at,
                }
            )

    return out


def _t2_detect_columns(matrix: list[list[Any]]) -> tuple[int, int | None, int | None, int | None] | None:
    for r in range(min(12, len(matrix))):
        row = matrix[r]
        if not row or row[0] is None:
            continue
        if not str(row[0]).strip().lower().startswith("commodity"):
            continue
        ncol = max(len(matrix[r + i]) for i in range(3) if r + i < len(matrix))
        col_text: list[str] = []
        for j in range(ncol):
            parts: list[str] = []
            for i in range(3):
                if r + i >= len(matrix) or j >= len(matrix[r + i]):
                    continue
                v = matrix[r + i][j]
                if v is not None and str(v).strip():
                    parts.append(str(v).lower())
            col_text.append(" ".join(parts))
        owner_j = loc_j = cap_j = None
        for j, t in enumerate(col_text):
            if "major operating" in t or ("equity" in t and "owner" in t):
                owner_j = j
            if "location" in t and "facilit" in t:
                loc_j = j
            if "capacity" in t or t.strip() == "annual":
                cap_j = j
        if owner_j is None or loc_j is None or cap_j is None:
            return None
        return r, owner_j, loc_j, cap_j
    return None


def _capacity_to_numeric(raw: str) -> float | None:
    s = _norm_text(raw).upper().replace(",", "")
    if not s or s in ("NA", "N/A", "--", "…"):
        return None
    s = s.rstrip(".")
    if re.fullmatch(r"-?\d+(\.\d+)?", s):
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _is_ditto_token(s: str) -> bool:
    """USGS 'Do.' / 'do.' in any column means ditto previous explicit value in that column."""
    t = _norm_text(s).lower().rstrip(".")
    return t == "do"


def _is_table2_wrap_continued_cell(c0: str) -> bool:
    """True if col A is only a wrap marker: 'Continued', '—Continued', etc.

    Lines like ``Iron and steel—Continued:`` end with ``:`` and are section headers, not wraps.
    """
    s = _norm_text(c0).strip()
    if not s or s.endswith(":"):
        return False
    return bool(
        re.fullmatch(r"[\u2014\u2013\u2012\-–]?\s*continued", s, re.IGNORECASE)
    )


def _is_t2_chrome_row(c0: str) -> bool:
    """Repeated headers and page chrome on Table 2 continuation pages — not facility rows."""
    low = _norm_text(c0).lower()
    if not low:
        return False
    if low in ("commodity", "commodity1", "commodity2"):
        return True
    if "see footnotes" in low:
        return True
    if low.startswith("table 2") and "continued" in low:
        return True
    if "structure of the mineral industry" in low and re.search(r"\b20\d{2}\b", low):
        return True
    return False


def _is_t2_header_fragment_row(ro: str, rl: str, rcap: str) -> bool:
    """Second-page subheader rows: empty commodity but cells repeat 'Major operating…' / 'Annual capacity'."""
    blob = " ".join((_norm_text(ro) + " " + _norm_text(rl) + " " + _norm_text(rcap)).lower().split())
    if "major operating" in blob or "equity owners" in blob:
        return True
    if "location of main" in blob and "facilit" in blob:
        return True
    if re.search(r"\bannual\b", blob) and "capacity" in blob:
        return True
    return False


# Table 2 — first word of a commodity cell that starts a new facility subtree (not a subtype line).
_T2_ROOT_FIRST_WORDS = frozenset(
    {
        "alumina",
        "aluminium",
        "aluminum",
        "ammonia",
        "bauxite",
        "cement",
        "chromium",
        "copper",
        "fertilizer",
        "gold",
        "gypsum",
        "iron",
        "lead",
        "lime",
        "manganese",
        "methanol",
        "nickel",
        "nitrogen",
        "petroleum",
        "phosphate",
        "potash",
        "salt",
        "sand",
        "silver",
        "stone",
        "sulfur",
        "urea",
        "zinc",
    }
)


def _t2_first_word(c0: str) -> str:
    parts = _norm_text(c0).split(",")[0].strip().split()
    if not parts:
        return ""
    return parts[0].lower().rstrip(":")


def _t2_colon_starts_new_branch(c0: str) -> bool:
    low = _norm_text(c0).lower().rstrip(":")
    return any(low.startswith(p) for p in ("iron and steel", "aluminum", "aluminium", "cement", "copper", "gold", "petroleum", "natural gas", "bauxite", "nickel", "zinc", "lead", "manganese", "chromium", "fertilizer", "nitrogen", "phosphate", "potash", "methanol", "sulfur", "salt", "sand", "stone", "gypsum", "lime", "alumina"))


def _myb3_fp_facility(
    source_file: str,
    country_iso3: str,
    facility_path: str,
    commodity_resolved: str,
    owner: str,
    loc: str,
    cap_raw: str,
    row_start: int,
    row_end: int,
) -> str:
    blob = "\x1e".join(
        [
            source_file,
            country_iso3,
            facility_path,
            commodity_resolved,
            owner,
            loc,
            cap_raw,
            str(row_start),
            str(row_end),
        ]
    )
    return hashlib.sha256(blob.encode("utf-8")).hexdigest()


def parse_myb3_table2(
    matrix: list[list[Any]],
    *,
    source_file: str,
    sheet_name: str,
    country_iso3: str,
    reference_year: int,
    pulled_at: str,
) -> list[dict[str, Any]]:
    """Parse Minerals Yearbook Table 2 into merged facility rows (see module NOTES for ditto / Continued / chrome)."""
    det = _t2_detect_columns(matrix)
    if not det:
        return []
    header_row, io, il, icap = det
    assert io is not None and il is not None and icap is not None

    unit_note: str | None = None
    for r in range(header_row):
        if r < len(matrix) and matrix[r] and matrix[r][0]:
            t = _norm_text(matrix[r][0])
            if "metric" in t.lower() or "ton" in t.lower():
                unit_note = t

    def cell(row: list[Any], idx: int | None) -> str:
        if idx is None or idx >= len(row):
            return ""
        v = row[idx]
        return _norm_text(v)

    facility_stack: list[str] = []
    last_non_do_commodity = ""
    last_owner = ""
    last_location = ""
    last_capacity = ""
    blocks: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None

    def flush() -> None:
        nonlocal current, last_owner, last_location, last_capacity
        if not current:
            return
        o = current["owner_operator"].strip()
        l = current["location"].strip()
        cr = current["capacity_raw"].strip()
        if not o and not l and not cr:
            current = None
            return
        rs, re_ = current["excel_row_start"], current["excel_row_end"]
        fp = _myb3_fp_facility(
            source_file,
            country_iso3,
            current["facility_path"],
            current["commodity_leaf_resolved"],
            o,
            l,
            cr,
            rs,
            re_,
        )
        blocks.append(
            {
                "record_fingerprint": fp,
                "country_iso3": country_iso3,
                "reference_year": reference_year,
                "commodity_cell_raw": current["commodity_cell_raw"],
                "commodity_leaf_resolved": current["commodity_leaf_resolved"],
                "facility_path": current["facility_path"],
                "owner_operator": o if o else None,
                "location": l if l else None,
                "capacity_raw": cr if cr else None,
                "capacity_numeric": _capacity_to_numeric(cr),
                "unit_note": unit_note,
                "sheet_name": sheet_name,
                "excel_row_start": rs,
                "excel_row_end": re_,
                "source_file": source_file,
                "source": SCRIPT_NAME,
                "pulled_at": pulled_at,
            }
        )
        # Ditto in the next row must repeat the full merged O/L/C, not only the first line of the block.
        if o:
            last_owner = o
        if l:
            last_location = l
        if cr:
            last_capacity = cr
        current = None

    for ri in range(header_row + 1, len(matrix)):
        row = matrix[ri]
        if not row:
            continue
        c0 = cell(row, 0)
        low = c0.lower()
        if low.startswith("eestimated") or ("ditto" in low and "zero" in low):
            break

        if _is_t2_chrome_row(c0):
            continue

        # Wrap-only col A: same commodity / block as the row above (incl. em-dash '—Continued').
        if _is_table2_wrap_continued_cell(c0):
            c0 = ""

        ro = cell(row, io)
        rl = cell(row, il)
        rcap = cell(row, icap)
        # Iran MYB3: stray 'content' in A with location-only continuation under chromite blocks.
        if current and _norm_text(c0).lower() == "content" and not ro and rl:
            c0 = ""

        has_olc = bool(ro or rl or rcap)

        if not _norm_text(c0) and has_olc and _is_t2_header_fragment_row(ro, rl, rcap):
            continue

        if c0.endswith(":") and not has_olc:
            flush()
            if _t2_colon_starts_new_branch(c0):
                facility_stack.clear()
            facility_stack.append(c0)
            continue

        if not c0 and not has_olc:
            continue

        if not c0 and has_olc and current:
            if ro and not _is_ditto_token(ro):
                current["owner_operator"] = (current["owner_operator"] + " " + ro).strip()
            if rl and not _is_ditto_token(rl):
                current["location"] = (current["location"] + " " + rl).strip()
            if rcap and not current["capacity_raw"]:
                cap_cont = last_capacity if _is_ditto_token(rcap) else rcap
                if cap_cont:
                    current["capacity_raw"] = cap_cont
            current["excel_row_end"] = ri
            continue

        if not has_olc:
            continue

        flush()

        if c0 and not _is_ditto_token(c0) and not c0.endswith(":"):
            fw = _t2_first_word(c0)
            if fw in _T2_ROOT_FIRST_WORDS:
                facility_stack.clear()

        raw_c = c0
        if _is_ditto_token(c0):
            leaf = last_non_do_commodity
        elif c0:
            last_non_do_commodity = c0
            leaf = c0
        else:
            leaf = last_non_do_commodity

        o = last_owner if _is_ditto_token(ro) else ro
        l = last_location if _is_ditto_token(rl) else rl
        cap = last_capacity if _is_ditto_token(rcap) else rcap

        if ro and not _is_ditto_token(ro):
            last_owner = ro
        if rl and not _is_ditto_token(rl):
            last_location = rl
        if rcap and not _is_ditto_token(rcap):
            last_capacity = rcap

        # Always persist a readable commodity label: ditto → resolved leaf; empty A with context → same leaf.
        resolved_leaf = leaf or raw_c or ""
        # facility_path: section stack + leaf (stack may be empty after root-word reset; leaf still labels row).
        stack_norm = [s.rstrip(":").strip() for s in facility_stack if _norm_text(s)]
        path_parts = stack_norm[:]
        if resolved_leaf and (not path_parts or path_parts[-1] != resolved_leaf):
            path_parts.append(resolved_leaf.strip())
        fpath = " > ".join(path_parts)
        if _is_ditto_token(raw_c):
            cell_raw_out = resolved_leaf or None
        elif raw_c:
            cell_raw_out = raw_c
        elif resolved_leaf:
            cell_raw_out = resolved_leaf
        else:
            cell_raw_out = None
        current = {
            "commodity_cell_raw": cell_raw_out,
            "commodity_leaf_resolved": resolved_leaf,
            "facility_path": fpath,
            "owner_operator": o,
            "location": l,
            "capacity_raw": cap,
            "excel_row_start": ri,
            "excel_row_end": ri,
        }

    flush()
    return blocks


def cmd_mcs(client: Any, run_id: int, file_arg: str | None) -> int:
    paths = _discover_mcs_files()
    if file_arg:
        p = USGS_DIR / file_arg
        if not p.is_file():
            msg = f"MCS file not found: {p}"
            finish_run(client, run_id, 0, "error", msg)
            print(msg, file=sys.stderr)
            return 1
        paths = [p]

    if not paths:
        msg = (
            f"No MCS CSV matching {MCS_GLOB} in {USGS_DIR}. "
            "Download MCS commodities data from USGS NMIC."
        )
        finish_run(client, run_id, 0, "error", msg)
        print(msg, file=sys.stderr)
        return 1

    pulled_at = datetime.now(timezone.utc).isoformat()
    total = 0
    try:
        for path in paths:
            try:
                df = pd.read_csv(path, encoding="cp1252", dtype=str, keep_default_na=False)
            except UnicodeDecodeError:
                df = pd.read_csv(path, encoding="latin-1", dtype=str, keep_default_na=False)
            rows = _mcs_dataframe_to_rows(df, path.name, pulled_at)
            if not rows:
                finish_run(client, run_id, 0, "partial", f"No rows parsed from {path.name}")
                print(f"No rows parsed from {path.name}", file=sys.stderr)
                return 1

            for i in range(0, len(rows), UPSERT_BATCH):
                batch = rows[i : i + UPSERT_BATCH]
                client.table("usgs_mineral_statistics").upsert(
                    batch, on_conflict="record_fingerprint"
                ).execute()
                total += len(batch)

        finish_run(client, run_id, total, "success", None)
        print(f"Upserted {total} MCS rows from {len(paths)} file(s).")
        return 0
    except Exception as e:
        err = f"{type(e).__name__}: {e}"
        finish_run(client, run_id, total, "error", err)
        print(err, file=sys.stderr)
        return 1


def cmd_facilities(client: Any, run_id: int) -> int:
    glob_paths = sorted(USGS_DIR.glob("myb3*.xlsx")) if USGS_DIR.is_dir() else []
    if not glob_paths:
        msg = f"No myb3*.xlsx under {USGS_DIR}."
        finish_run(client, run_id, 0, "error", msg)
        print(msg, file=sys.stderr)
        return 1

    pulled_at = datetime.now(timezone.utc).isoformat()
    prod_rows: list[dict[str, Any]] = []
    fac_rows: list[dict[str, Any]] = []
    errors: list[str] = []

    for path in glob_paths:
        meta = _parse_myb3_filename(path.name)
        if not meta:
            errors.append(f"Skip unrecognized myb3 name: {path.name}")
            continue
        ref_year, iso3 = meta

        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as e:
            errors.append(f"{path.name}: {e}")
            continue

        try:
            s1 = _find_workbook_sheet(wb.sheetnames, r"table\s*1")
            s2 = _find_workbook_sheet(wb.sheetnames, r"table\s*2")
            if not s1 or not s2:
                errors.append(f"{path.name}: missing Table 1 or Table 2 sheet")
                continue

            m1 = _sheet_to_matrix(wb[s1])
            m2 = _sheet_to_matrix(wb[s2])

            prod_rows.extend(
                parse_myb3_table1(
                    m1,
                    source_file=path.name,
                    sheet_name=s1,
                    country_iso3=iso3,
                    reference_year=ref_year,
                    pulled_at=pulled_at,
                )
            )
            fac_rows.extend(
                parse_myb3_table2(
                    m2,
                    source_file=path.name,
                    sheet_name=s2,
                    country_iso3=iso3,
                    reference_year=ref_year,
                    pulled_at=pulled_at,
                )
            )
        finally:
            wb.close()

    if errors and not prod_rows and not fac_rows:
        finish_run(client, run_id, 0, "error", "; ".join(errors))
        print("; ".join(errors), file=sys.stderr)
        return 1

    # Remove prior rows for each workbook we are about to load. Upsert keys on
    # record_fingerprint alone leave orphans when fingerprints change (parser fixes).
    cleared: list[str] = []
    for path in glob_paths:
        if not _parse_myb3_filename(path.name):
            continue
        fn = path.name
        client.table("usgs_myb3_production").delete().eq("source_file", fn).execute()
        client.table("usgs_country_mineral_facilities").delete().eq("source_file", fn).execute()
        cleared.append(fn)

    total = 0
    try:
        for i in range(0, len(prod_rows), UPSERT_BATCH):
            batch = prod_rows[i : i + UPSERT_BATCH]
            client.table("usgs_myb3_production").upsert(
                batch, on_conflict="record_fingerprint"
            ).execute()
            total += len(batch)
        for i in range(0, len(fac_rows), UPSERT_BATCH):
            batch = fac_rows[i : i + UPSERT_BATCH]
            client.table("usgs_country_mineral_facilities").upsert(
                batch, on_conflict="record_fingerprint"
            ).execute()
            total += len(batch)

        msg_extra = f" ({'; '.join(errors)})" if errors else ""
        status = "success" if not errors else "partial"
        finish_run(
            client,
            run_id,
            total,
            status,
            "; ".join(errors) if errors else None,
        )
        print(
            f"Cleared prior rows for {len(cleared)} workbook(s); "
            f"upserted {len(prod_rows)} myb3 production + {len(fac_rows)} facility rows "
            f"from {len(glob_paths)} file(s).{msg_extra}"
        )
        return 0
    except Exception as e:
        err = f"{type(e).__name__}: {e}"
        finish_run(client, run_id, total, "error", err)
        print(err, file=sys.stderr)
        return 1


def main() -> int:
    ap = argparse.ArgumentParser(description="Load USGS files into Supabase.")
    ap.add_argument(
        "command",
        nargs="?",
        default="mcs",
        choices=("mcs", "facilities"),
        help="mcs = MCS commodities CSV (default); facilities = myb3 xlsx (Table 1 + Table 2)",
    )
    ap.add_argument(
        "--file",
        type=str,
        help=f"MCS filename under data/usgs/ (mcs only; default: all {MCS_GLOB})",
    )

    args = ap.parse_args()

    try:
        client = get_client()
    except RuntimeError as e:
        print(e, file=sys.stderr)
        return 1

    params: dict[str, Any] = {"command": args.command, "file": args.file}

    try:
        src = SOURCE_LABEL_MYB3 if args.command == "facilities" else SOURCE_LABEL
        run_id = start_run(client, SCRIPT_NAME, src, params)
    except Exception as e:
        print(f"Could not log pipeline run (check Supabase keys): {e}", file=sys.stderr)
        return 1

    if args.command == "mcs":
        return cmd_mcs(client, run_id, args.file)
    if args.command == "facilities":
        return cmd_facilities(client, run_id)
    finish_run(client, run_id, 0, "error", f"Unknown command {args.command}")
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
